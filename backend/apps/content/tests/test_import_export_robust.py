import pytest
import io
from django.utils.text import slugify
from apps.content.models import Book, Author, Category
from apps.content.utils.import_export import BookImportExport
from unittest.mock import patch
from datetime import date
import openpyxl

@pytest.mark.django_db
class TestBookImportExportRobust:
    """Robust tests for book import/export utility"""

    def test_export_to_csv(self):
        """Test exporting books to CSV"""
        author = Author.objects.create(name="Author CSV")
        category = Category.objects.create(name="Category CSV", slug="cat-csv")
        Book.objects.create(
            title="Book CSV",
            author=author,
            category=category,
            description="Desc CSV",
            isbn="1234567890123",
            is_premium=True,
            slug="book-csv"
        )
        
        queryset = Book.objects.filter(title="Book CSV")
        content = BookImportExport.export_books(queryset, 'csv').decode('utf-8')
        
        assert "Book CSV" in content
        assert "Author CSV" in content
        assert "Category CSV" in content
        assert "SÍ" in content

    def test_export_to_xlsx(self):
        """Test exporting books to XLSX"""
        author = Author.objects.create(name="Author XLSX")
        Book.objects.create(
            title="Book XLSX",
            author=author,
            description="Desc XLSX",
            slug="book-xlsx"
        )
        
        queryset = Book.objects.filter(title="Book XLSX")
        content = BookImportExport.export_books(queryset, 'xlsx')
        
        # Verify it's a valid XLSX by loading it
        f = io.BytesIO(content)
        wb = openpyxl.load_workbook(f)
        ws = wb.active
        assert ws.cell(row=2, column=1).value == "Book XLSX"

    def test_import_from_csv(self):
        """Test importing books from CSV"""
        csv_content = (
            "Título,Autor,Categoría,Descripción,ISBN,Es Premium\n"
            "New Book CSV,New Author,New Category,Some description,1112223334445,sí\n"
        )
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        
        result = BookImportExport.import_books(file_obj, 'csv')
        
        assert result['imported'] == 1
        assert Book.objects.filter(title="New Book CSV").exists()
        book = Book.objects.get(title="New Book CSV")
        assert book.author.name == "New Author"
        assert book.category.name == "New Category"
        assert book.is_premium is True

    def test_import_duplicate_skipped(self):
        """Test that existing books are skipped"""
        author = Author.objects.create(name="Existing Author")
        Book.objects.create(title="Existing Book", author=author, slug="existing-book")
        
        csv_content = (
            "Título,Autor,Categoría,Descripción,ISBN,Es Premium\n"
            "Existing Book,Existing Author,Category,Desc,123,no\n"
        )
        file_obj = io.BytesIO(csv_content.encode('utf-8'))
        
        result = BookImportExport.import_books(file_obj, 'csv')
        
        assert result['imported'] == 0
        assert result['skipped'] == 1

    def test_import_from_xlsx(self):
        """Test importing books from XLSX"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Título", "Autor", "Categoría", "Descripción", "ISBN", "Es Premium"])
        ws.append(["XLSX Book", "XLSX Author", "XLSX Category", "XLSX Desc", "9998887776665", "SÍ"])
        
        f = io.BytesIO()
        wb.save(f)
        f.seek(0)
        
        result = BookImportExport.import_books(f, 'xlsx')
        
        assert result['imported'] == 1
        assert Book.objects.filter(title="XLSX Book").exists()
        book = Book.objects.get(title="XLSX Book")
        assert book.is_premium is True
