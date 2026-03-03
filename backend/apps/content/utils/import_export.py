import csv
import io
import logging
from datetime import datetime
from django.utils.text import slugify
from django.core.files.base import ContentFile
from apps.content.models import Book, Author, Category
import openpyxl
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class BookImportExport:
    """
    Utility class for importing and exporting books in CSV and XLSX formats.
    """

    # Ordered list of (field_key, header_label, example_value, notes)
    FIELDS = [
        ('title',            'Título',                  'El principito',                  'Requerido'),
        ('author',           'Autor',                   'Antoine de Saint-Exupéry',       'Nombre completo'),
        ('category',         'Categoría',               'Literatura infantil',             ''),
        ('description',      'Descripción',             'Historia de un pequeño príncipe que viaja por planetas.', ''),
        ('isbn',             'ISBN',                    '978-0156013987',                  'Máx. 13 caracteres'),
        ('published_year',   'Año de Publicación',      '1943',                            'Solo el año (número)'),
        ('publication_date', 'Fecha de Publicación',    '1943-04-06',                      'Formato AAAA-MM-DD'),
        ('publisher',        'Editorial',               'Reynal & Hitchcock',              ''),
        ('language',         'Idioma',                  'es',                              'Código ISO: es, en, pt, fr…'),
        ('is_premium',       'Es Premium',              'NO',                              'SÍ o NO'),
        ('is_open_access',   'Acceso Abierto',          'NO',                              'SÍ o NO'),
        ('source',           'Fuente',                  'manual',                          'manual, openlibrary o doab'),
        ('doi',              'DOI',                     '',                                'Ej: 10.xxxx/xxxxx'),
        ('external_url',     'URL Externa',             '',                                'URL del PDF externo (libros OA)'),
        ('cover_image',      'Portada (URL)',            '',                                'URL de imagen de portada'),
    ]

    # Mapping for import: header → field_key
    @classmethod
    def _reverse_mapping(cls):
        return {label.lower(): key for key, label, *_ in cls.FIELDS}

    # -------------------------------------------------------------------------
    # EXPORT
    # -------------------------------------------------------------------------

    @classmethod
    def export_books(cls, queryset, format_type='csv', request=None):
        if format_type == 'csv':
            return cls._export_to_csv(queryset, request)
        elif format_type == 'xlsx':
            return cls._export_to_xlsx(queryset, request)
        raise ValueError(f"Unsupported format: {format_type}")

    @classmethod
    def _book_to_row(cls, book, request=None):
        cover = ''
        if book.cover_image:
            cover = request.build_absolute_uri(book.cover_image.url) if request else book.cover_image.url

        return {
            'Título': book.title,
            'Autor': book.author.name if book.author else '',
            'Categoría': book.category.name if book.category else '',
            'Descripción': book.description or '',
            'ISBN': book.isbn or '',
            'Año de Publicación': book.published_year or '',
            'Fecha de Publicación': book.publication_date.strftime('%Y-%m-%d') if book.publication_date else '',
            'Editorial': book.publisher or '',
            'Idioma': book.language or '',
            'Es Premium': 'SÍ' if book.is_premium else 'NO',
            'Acceso Abierto': 'SÍ' if book.is_open_access else 'NO',
            'Fuente': book.source or 'manual',
            'DOI': book.doi or '',
            'URL Externa': book.external_url or '',
            'Portada (URL)': cover,
        }

    @classmethod
    def _export_to_csv(cls, queryset, request=None):
        headers = [label for _, label, *_ in cls.FIELDS]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        for book in queryset:
            writer.writerow(cls._book_to_row(book, request))
        return output.getvalue().encode('utf-8-sig')

    @classmethod
    def _export_to_xlsx(cls, queryset, request=None):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Libros"
        headers = [label for _, label, *_ in cls.FIELDS]

        header_font = Font(bold=True)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font

        for row_num, book in enumerate(queryset, 2):
            row = cls._book_to_row(book, request)
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row.get(header, ''))

        cls._autofit_columns(ws)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # -------------------------------------------------------------------------
    # IMPORT TEMPLATE
    # -------------------------------------------------------------------------

    @classmethod
    def generate_template(cls, format_type='csv'):
        """Return a file with headers + one populated example row."""
        headers = [label for _, label, *_ in cls.FIELDS]
        example = [example for _, _, example, *_ in cls.FIELDS]
        notes = [note for _, _, _, note, *_ in cls.FIELDS] if len(cls.FIELDS[0]) == 4 else []

        if format_type == 'csv':
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerow(example)
            return output.getvalue().encode('utf-8-sig')

        # XLSX
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Plantilla"

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        example_fill = PatternFill(start_color='EBF4FF', end_color='EBF4FF', fill_type='solid')
        notes_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
        notes_font = Font(italic=True, color='888888', size=9)

        for col_num, (header, example_val) in enumerate(zip(headers, example), 1):
            # Row 1: header
            hcell = ws.cell(row=1, column=col_num, value=header)
            hcell.font = header_font
            hcell.fill = header_fill
            hcell.alignment = Alignment(horizontal='center')

            # Row 2: example
            ecell = ws.cell(row=2, column=col_num, value=example_val)
            ecell.fill = example_fill

        # Row 3: notes (if any)
        if notes:
            for col_num, note in enumerate(notes, 1):
                ncell = ws.cell(row=3, column=col_num, value=note)
                ncell.fill = notes_fill
                ncell.font = notes_font

        cls._autofit_columns(ws)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    # -------------------------------------------------------------------------
    # IMPORT
    # -------------------------------------------------------------------------

    @classmethod
    def import_books(cls, file_obj, format_type):
        if format_type == 'csv':
            data = cls._parse_csv(file_obj)
        elif format_type == 'xlsx':
            data = cls._parse_xlsx(file_obj)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
        return cls._process_import(data)

    @classmethod
    def _parse_csv(cls, file_obj):
        try:
            decoded_file = file_obj.read().decode('utf-8-sig')
        except UnicodeDecodeError:
            file_obj.seek(0)
            decoded_file = file_obj.read().decode('latin-1')

        io_string = io.StringIO(decoded_file)
        reader = csv.DictReader(io_string)
        reverse = cls._reverse_mapping()

        data = []
        for row in reader:
            normalized = {}
            for k, v in row.items():
                if not k:
                    continue
                key = reverse.get(k.strip().lower(), k.strip().lower())
                normalized[key] = v
            data.append(normalized)
        return data

    @classmethod
    def _parse_xlsx(cls, file_obj):
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        reverse = cls._reverse_mapping()

        headers = [cell.value for cell in ws[1]]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            normalized = {}
            for idx, value in enumerate(row):
                if idx < len(headers) and headers[idx]:
                    header = str(headers[idx]).strip().lower()
                    key = reverse.get(header, header)
                    normalized[key] = value
            data.append(normalized)
        return data

    @classmethod
    def _process_import(cls, data):
        imported_count = 0
        skipped_count = 0
        errors = []

        for row in data:
            # Skip the notes row (row 3 in template) if it got parsed
            title = row.get('title')
            if not title or str(title).strip().lower() in ('requerido', 'título', ''):
                continue

            try:
                if not str(title).strip():
                    errors.append("Fila omitida: Título faltante.")
                    continue

                book_defaults = {
                    'description': str(row.get('description', '') or ''),
                    'isbn': str(row.get('isbn', '') or '')[:13],
                    'is_premium': str(row.get('is_premium', '') or '').lower() in ['sì', 'si', 'sí', 'true', '1', 'yes'],
                    'is_open_access': str(row.get('is_open_access', '') or '').lower() in ['sì', 'si', 'sí', 'true', '1', 'yes'],
                    'publisher': str(row.get('publisher', '') or ''),
                    'language': str(row.get('language', '') or ''),
                    'doi': str(row.get('doi', '') or '') or None,
                    'external_url': str(row.get('external_url', '') or '') or None,
                }

                source = str(row.get('source', '') or '').lower()
                if source in ('manual', 'openlibrary', 'doab'):
                    book_defaults['source'] = source

                # Published year
                py = row.get('published_year')
                if py:
                    try:
                        book_defaults['published_year'] = int(str(py))
                    except (ValueError, TypeError):
                        pass

                # Author
                author_name = str(row.get('author', '') or '').strip()
                if author_name:
                    author, _ = Author.objects.get_or_create(name=author_name)
                    book_defaults['author'] = author

                # Category
                category_name = str(row.get('category', '') or '').strip()
                if category_name:
                    category, _ = Category.objects.get_or_create(
                        name=category_name,
                        defaults={'slug': slugify(category_name)},
                    )
                    book_defaults['category'] = category

                # Publication date
                pub_date = row.get('publication_date')
                if pub_date:
                    if isinstance(pub_date, datetime):
                        book_defaults['publication_date'] = pub_date.date()
                    else:
                        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y'):
                            try:
                                dt = datetime.strptime(str(pub_date), fmt)
                                book_defaults['publication_date'] = dt.date()
                                break
                            except ValueError:
                                continue

                # Slug + duplicate check
                slug = row.get('slug') or slugify(str(title))
                if Book.objects.filter(slug=slug).exists():
                    skipped_count += 1
                    continue

                # Cover image
                cover_url = row.get('cover_image')
                cover_file = None
                if cover_url and str(cover_url).startswith(('http://', 'https://')):
                    cover_file = cls._download_image(str(cover_url))

                book = Book.objects.create(title=str(title), slug=slug, **book_defaults)
                if cover_file:
                    book.cover_image.save(f'cover_{book.id}.jpg', cover_file, save=True)

                imported_count += 1

            except Exception as e:
                logger.error(f"Error importing row: {row}. Error: {e}")
                errors.append(f"Error en '{row.get('title', 'Desconocido')}': {e}")

        return {'imported': imported_count, 'skipped': skipped_count, 'errors': errors}

    # -------------------------------------------------------------------------
    # HELPERS
    # -------------------------------------------------------------------------

    @staticmethod
    def _autofit_columns(ws):
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value and len(str(cell.value)) > max_len:
                        max_len = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 4, 55)

    @staticmethod
    def _download_image(url):
        try:
            response = requests.get(url, timeout=10)
            if response.status_code == 200:
                return ContentFile(response.content)
            logger.error(f"Failed to download image from {url}: {response.status_code}")
        except Exception as e:
            logger.error(f"Exception downloading image from {url}: {e}")
        return None
